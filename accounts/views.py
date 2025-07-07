# views.py
from django.views.generic import CreateView, UpdateView, ListView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from pdfkit import pdfkit

from .models import Company, Document, AuditLog, CustomUser, SupportTicket
from .forms import CompanyRegistrationForm, DocumentUploadForm, DeliveryAddressForm
import random
import string
from django.core.mail import send_mail
from django.contrib.auth import login
from django.urls import reverse, reverse_lazy
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.utils import timezone
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.conf import settings
#from xhtml2pdf import pisa
from openpyxl import Workbook
from io import BytesIO
import datetime
from django.core.files.base import ContentFile
from .forms import CartItemForm, OrderCreateForm
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views.generic import ListView,  TemplateView
from main.models import Order, OrderItem, Invoice, Cart, CartItem, DeliveryAddress
from .models import Company
from django.http import FileResponse

import logging
logger = logging.getLogger(__name__)


class CompanyRegisterView(CreateView):
    model = Company
    form_class = CompanyRegistrationForm
    template_name = 'registration/company_register.html'
    success_url = reverse_lazy('accounts:verify_email_sent')

    @transaction.atomic
    def form_valid(self, form):
        print("Начало обработки формы регистрации компании")

        try:
            # Сохраняем компанию с временным токеном подтверждения
            company = form.save(commit=False)
            company.organization_type = form.data.get('organization_type', 'ООО')
            company.company_status = form.data.get('company_status', 'Действующее')
            token = ''.join(random.choices(string.ascii_letters + string.digits, k=50))
            company.verification_token = token
            company.verification_token_created_at = timezone.now()  # Время создания токена
            company.is_verified = False

            print(f"Данные формы: {form.data}")
            print(f"Organization type: {form.data.get('organization_type')}")
            print(f"Company status: {form.data.get('company_status')}")
            print(f"Сохранение компании: {company.legal_name}")
            company.save()

            # Создаем администратора компании
            admin_user = CustomUser.objects.create(
                email=form.cleaned_data['email'],
                password=make_password(form.cleaned_data['password']),
                company=company,
                role=form.cleaned_data['role'],
                is_active=False,
                phone='',
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
                middle_name=form.cleaned_data.get('middle_name', '')  # Сохраняем отчество
            )
            print(f"Создан пользователь: {admin_user.email}")

            # Отправляем email с подтверждением
            verification_url = self.request.build_absolute_uri(
                reverse('accounts:verify_company_email', kwargs={'token': token})
            )
            print(f"Ссылка подтверждения: {verification_url}")

            subject = 'Подтверждение регистрации компании'
            message = (
                f'Для завершения регистрации вашей компании "{company.legal_name}" перейдите по ссылке:\n\n'
                f'{verification_url}\n\n'
                f'Ваши данные для входа:\n'
                f'Email: {admin_user.email}\n'
                f'Пароль: {form.cleaned_data["password"]}\n\n'
                'Ссылка действительна в течение 5 минут.\n\n'
                'После входа вам будет предложено сменить пароль.'
            )

            print(f"Отправка email на {admin_user.email}")
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [admin_user.email],
                fail_silently=False,
            )

            # Сохраняем ID компании в сессии
            self.request.session['new_company_id'] = company.id
            print(f"Сохранено в сессии: new_company_id={company.id}")

            print(f"Перенаправление на: {self.success_url}")
            return redirect(self.success_url)

        except Exception as e:
            print(f"Ошибка при обработке формы: {str(e)}")
            messages.error(self.request, f'Ошибка при регистрации: {str(e)}')
            return self.form_invalid(form)


def verify_company_email(request, token):
    try:
        company = Company.objects.get(verification_token=token)

        # Проверяем, не истекло ли время действия токена (5 минут)
        token_age = timezone.now() - company.verification_token_created_at
        if token_age > timedelta(minutes=5):
            messages.error(request, '❌ Срок действия ссылки истёк. Пожалуйста, запросите новую.')
            return redirect('accounts:invalid_token')

        # Подтверждаем компанию и активируем пользователя
        company.is_verified = True
        company.verification_token = None
        company.verification_token_created_at = None
        company.save()

        admin_user = CustomUser.objects.get(email=company.email)
        admin_user.is_active = True
        admin_user.save()

        # Авторизуем пользователя
        login(request, admin_user, backend='django.contrib.auth.backends.ModelBackend')

        return redirect('accounts:company_dashboard')

    except (Company.DoesNotExist, CustomUser.DoesNotExist):
        messages.error(request, '❌ Недействительная ссылка подтверждения')
        return redirect('accounts:invalid_token')

    except Exception as e:
        messages.error(request, f'❌ Ошибка: {str(e)}')
        return redirect('accounts:invalid_token')


class CompanyDashboardView(LoginRequiredMixin, ListView):
    template_name = 'accounts/company_dashboard.html'
    paginate_by = 10

    def get_queryset(self):
        return Order.objects.filter(company=self.request.user.company).prefetch_related('items').order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company
        user = self.request.user

        if not company:
            messages.error(self.request, "Ваш аккаунт не привязан к компании")
            return context

        # Получаем документы (счета) компании
        documents = Document.objects.filter(
            company=company,
            doc_type='invoice'
        ).select_related('invoice').order_by('-created_at')

        context.update({
            'company': company,
            'user': user,
            'orders': context['object_list'],
            'documents': documents,  # Добавляем документы в контекст
            'delivery_addresses': company.delivery_addresses.all() if company else [],
            'delivery_address_form': DeliveryAddressForm(),
            'activity': self._get_activity(company)
        })
        return context

    def _get_activity(self, company):
        orders = Order.objects.filter(company=company).select_related('invoice').order_by('-created_at')[:10]
        activity = []

        for order in orders:
            activity.append({
                'type': 'order_created',
                'object': order,
                'date': order.created_at,
                'message': f'Создан заказ #{order.id}',
                'status': order.get_status_display()
            })

            if hasattr(order, 'invoice'):
                # Добавляем проверку на существование файла
                if order.invoice.pdf_file:
                    activity.append({
                        'type': 'invoice_issued',
                        'object': order.invoice,
                        'date': order.invoice.created_at,
                        'message': f'Выставлен счет #{order.invoice.invoice_number}',
                        'status': 'Ожидает оплаты' if not order.invoice.paid else 'Оплачен'
                    })

            if order.status == Order.STATUS_IN_PROGRESS:
                activity.append({
                    'type': 'order_pending',
                    'object': order,
                    'date': order.updated_at,
                    'message': f'Заказ #{order.id} ожидает оплаты',
                    'status': 'Ожидает оплаты'
                })

        return sorted(activity, key=lambda x: x['date'], reverse=True)[:15]

def order_list(request):
    orders = Order.objects.filter(company=request.user.company).order_by('-created_at')
    return render(request, 'accounts/orders_list.html', {
        'orders': orders
    })

def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, company=request.user.company)
    return render(request, 'accounts/order_detail.html', {
        'order': order,
        'invoice': getattr(order, 'invoice', None)
    })




class InvalidTokenView(TemplateView):
    template_name = 'accounts/emails/invalid_token.html'


class DocumentUploadView(LoginRequiredMixin, CreateView):
    model = Document
    form_class = DocumentUploadForm
    template_name = 'accounts/upload_document.html'

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        return super().form_valid(form)


class OrderCreateView(LoginRequiredMixin, CreateView):
    model = Order
    form_class = OrderCreateForm
    template_name = 'dashboard/create_order.html'

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        return super().form_valid(form)


def get_user_cart(request):
    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user)
        return cart
    return None


def cart_view(request):
    cart = get_user_cart(request)
    if not cart:
        return redirect('accounts:login')

    if request.method == 'POST':
        form = CartItemForm(request.POST)
        if form.is_valid():
            item_id = request.POST.get('item_id')
            try:
                item = CartItem.objects.get(id=item_id, cart=cart)
                item.quantity = form.cleaned_data['quantity']
                item.save()
                messages.success(request, 'Количество товара обновлено')
            except CartItem.DoesNotExist:
                messages.error(request, 'Товар не найден в корзине')
        return redirect('accounts:cart_view')

    return render(request, 'accounts/cart.html', {
        'cart': cart,
        'cart_items': cart.items.all() if cart else []
    })


def add_to_cart(request, product_id):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Требуется авторизация'}, status=403)

    # Здесь должна быть логика получения данных о товаре из вашей системы
    # Это пример - замените на реальное получение данных о товаре
    product_data = {
        'id': product_id,
        'name': request.POST.get('product_name', 'Тестовый товар'),
        'price': float(request.POST.get('price', 1000)),
        'image': request.POST.get('image', '')
    }

    cart = get_user_cart(request)
    if not cart:
        return JsonResponse({'status': 'error', 'message': 'Ошибка корзины'}, status=400)

    # Проверяем, есть ли уже такой товар в корзине
    item, created = CartItem.objects.get_or_create(
        cart=cart,
        product_id=product_data['id'],
        defaults={
            'product_name': product_data['name'],
            'product_image': product_data['image'],
            'price': product_data['price'],
            'quantity': 1
        }
    )

    if not created:
        item.quantity += 1
        item.save()

    messages.success(request, 'Товар добавлен в корзину')
    return JsonResponse({
        'status': 'success',
        'cart_items_count': cart.items_count,
        'cart_total': cart.total_price
    })


def remove_from_cart(request, item_id):
    cart = get_user_cart(request)
    if not cart:
        return redirect('accounts:login')

    try:
        item = CartItem.objects.get(id=item_id, cart=cart)
        item.delete()
        messages.success(request, 'Товар удален из корзины')
    except CartItem.DoesNotExist:
        messages.error(request, 'Товар не найден в корзине')

    return redirect('accounts:cart_view')


def checkout(request):
    cart = get_user_cart(request)
    if not cart or cart.items_count == 0:
        return redirect('accounts:cart_view')

    if request.method == 'POST':
        form = OrderCreateForm(request.POST)
        if form.is_valid():
            # Создаем заказ с явным статусом "new"
            order = Order.objects.create(
                company=request.user.company,
                total_amount=cart.total_price,
                notes=form.cleaned_data['notes'],
                status='new'  # Явно устанавливаем статус
            )

            # Переносим товары из корзины в заказ
            for cart_item in cart.items.all():
                OrderItem.objects.create(
                    order=order,
                    product_id=cart_item.product_id,
                    product_name=cart_item.product_name,
                    product_image=cart_item.product_image,
                    price=cart_item.price,
                    quantity=cart_item.quantity,
                    total_price=cart_item.total_price
                )

            # Очищаем корзину
            cart.items.all().delete()

            # Создаем счет
            invoice = create_invoice(order)

            # Отправляем письмо с подтверждением
            send_order_confirmation(request, order, invoice)

            messages.success(request, 'Ваш заказ успешно оформлен!')
            return redirect('accounts:order_detail', order_id=order.id)
    else:
        form = OrderCreateForm()

    return render(request, 'accounts/checkout.html', {
        'cart': cart,
        'form': form,
        'company': request.user.company
    })


def create_invoice(order):
    """Создает счет и связанный документ"""
    try:
        # Генерация PDF
        pdf_content = generate_invoice_pdf(order)
        if not pdf_content:
            raise ValueError("Не удалось сгенерировать PDF")

        # Создание временного файла
        invoice_number = f"INV-{order.id}-{datetime.datetime.now().strftime('%Y%m%d')}"
        pdf_file = ContentFile(pdf_content, name=f'invoice_{invoice_number}.pdf')

        # Создаем документ для личного кабинета
        document = Document.objects.create(
            company=order.company,
            doc_type='invoice',
            file=pdf_file,
            signed=False
        )

        # Создаем счет
        invoice = Invoice.objects.create(
            order=order,
            invoice_number=invoice_number,
            due_date=datetime.datetime.now() + datetime.timedelta(days=7),
            amount=order.total_amount,
            pdf_file=pdf_file,
            document=document  # Связываем с документом
        )

        return invoice

    except Exception as e:
        logger.error(f"Ошибка при создании счета: {str(e)}", exc_info=True)
        raise


def generate_invoice_pdf(invoice):
    """Генерирует PDF счета"""
    context = {
        'invoice': invoice,
        'company': invoice.order.company,
        'order': invoice.order,
        'items': invoice.order.items.all()
    }

    html = render_to_string('invoice_pdf.html', context)
    result = BytesIO()

    try:
        # Укажите правильный путь к wkhtmltopdf
        config = pdfkit.configuration(
            wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'  # Или другой путь
        )
        pdf = pdfkit.from_string(html, False, configuration=config)
        return pdf
    except Exception as e:
        logger.error(f"Ошибка генерации PDF: {e}")
        return None

def generate_invoice_excel(invoice):
    wb = Workbook()
    ws = wb.active
    ws.title = "Счет"

    # Заголовок
    ws['A1'] = f"Счет № {invoice.invoice_number}"
    ws['A2'] = f"Дата: {invoice.created_at.strftime('%d.%m.%Y')}"
    ws['A3'] = f"Срок оплаты: {invoice.due_date.strftime('%d.%m.%Y')}"
    ws['A4'] = f"Заказчик: {invoice.order.company.legal_name}"
    ws['A5'] = f"ИНН: {invoice.order.company.inn}"

    # Заголовки таблицы
    ws['A7'] = "№"
    ws['B7'] = "Наименование товара"
    ws['C7'] = "Количество"
    ws['D7'] = "Цена"
    ws['E7'] = "Сумма"

    # Данные товаров
    for i, item in enumerate(invoice.order.items.all(), start=1):
        ws[f'A{7 + i}'] = i
        ws[f'B{7 + i}'] = item.product_name
        ws[f'C{7 + i}'] = item.quantity
        ws[f'D{7 + i}'] = float(item.price)
        ws[f'E{7 + i}'] = float(item.total_price)

    # Итого
    last_row = 7 + invoice.order.items.count()
    ws[f'D{last_row + 1}'] = "Итого:"
    ws[f'E{last_row + 1}'] = float(invoice.amount)

    # Сохранение в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    invoice.excel_file.save(
        f'invoice_{invoice.invoice_number}.xlsx',
        excel_file
    )
    invoice.save()


def send_order_confirmation(request, order, invoice):
    subject = f"Подтверждение заказа #{order.id}"
    context = {
        'order': order,
        'invoice': invoice,
        'company': request.user.company,
        'user': request.user,
        'site_url': request.build_absolute_uri('/')
    }

    html_message = render_to_string('accounts/emails/order_confirmation.html', context)
    text_message = render_to_string('accounts/emails/order_confirmation.txt', context)

    email = EmailMessage(
        subject,
        text_message,
        settings.DEFAULT_FROM_EMAIL,
        [request.user.email, order.company.email],
        reply_to=[settings.DEFAULT_FROM_EMAIL]
    )
    email.attach_alternative(html_message, "text/html")

    # Прикрепляем PDF счета
    if invoice.pdf_file:
        email.attach_file(invoice.pdf_file.path)

    # Прикрепляем Excel счета
    if invoice.excel_file:
        email.attach_file(invoice.excel_file.path)

    email.send()


def order_list(request):
    orders = Order.objects.filter(company=request.user.company).order_by('-created_at')
    return render(request, 'accounts/orders_list.html', {
        'orders': orders
    })


def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, company=request.user.company)
    return render(request, 'accounts/order_detail.html', {
        'order': order,
        'invoice': getattr(order, 'invoice', None)
    })
class CompanyProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = Company
    fields = ['legal_address', 'bank_account', 'bank_bik']
    template_name = 'accounts/company_dashboard.html'
    success_url = reverse_lazy('accounts:company_dashboard')

    def get_object(self):
        return self.request.user.company

    def form_valid(self, form):
        form.save()
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'legal_address': form.instance.legal_address or '',
                'bank_account': form.instance.bank_account or '',
                'bank_bik': form.instance.bank_bik or ''
            })
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'Неверные данные формы',
                'errors': form.errors
            }, status=400)
        return super().form_invalid(form)

def resend_verification(request, company_id):
    try:
        company = get_object_or_404(Company, id=company_id)

        # Генерируем новый токен и обновляем время
        new_token = ''.join(random.choices(string.ascii_letters + string.digits, k=50))
        company.verification_token = new_token
        company.verification_token_created_at = timezone.now()
        company.save()

        verification_url = request.build_absolute_uri(
            reverse('accounts:verify_company_email', kwargs={'token': new_token})
        )

        subject = 'Повторное подтверждение регистрации компании'
        message = (
            f'Для завершения регистрации вашей компании "{company.legal_name}" перейдите по ссылке:\n\n'
            f'{verification_url}\n\n'
            f'Ссылка действительна в течение 5 минут.\n\n'
            f'Если вы не запрашивали повторную отправку, проигнорируйте это письмо.'
        )

        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [company.email],
            fail_silently=False,
        )

        messages.success(request, '✅ Письмо с подтверждением отправлено повторно')
        return redirect('accounts:verify_email_sent')

    except Exception as e:
        messages.error(request, f'❌ Ошибка: {str(e)}')
        return redirect('accounts:invalid_token')


def check_verification_status(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    return JsonResponse({
        'is_verified': company.is_verified,
        'company_name': company.legal_name
    })

def download_document(request, pk):
    """Скачивание документа"""
    document = get_object_or_404(
        Document,
        pk=pk,
        company=request.user.company
    )
    response = FileResponse(document.file.open('rb'))
    response['Content-Disposition'] = f'attachment; filename="{document.file.name}"'
    return response

class EmailVerificationSentView(TemplateView):
    template_name = 'registration/verify_email_sent.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company_id = self.request.session.get('new_company_id')
        if company_id:
            company = get_object_or_404(Company, id=company_id)
            context['company'] = company
            context['expiration_minutes'] = 5  # Время действия ссылки

            if company.verification_token:
                context['verification_link'] = self.request.build_absolute_uri(
                    reverse('accounts:verify_company_email', kwargs={'token': company.verification_token})
                )
        return context


class FinancialDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/financial_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company

        context.update({
            'company': company,
            'total_orders_amount': 1245000,
            'average_order_amount': 24900,
            'unpaid_invoices': 245000,
        })
        return context


class OrdersDashboardView(LoginRequiredMixin, ListView):
    template_name = 'accounts/orders_dashboard.html'
    context_object_name = 'orders'

    def get_queryset(self):
        return Order.objects.filter(company=self.request.user.company).order_by('-created_at')[:10]


class DocumentsDashboardView(LoginRequiredMixin, ListView):
    template_name = 'accounts/documents_dashboard.html'
    context_object_name = 'documents'

    def get_queryset(self):
        return Document.objects.filter(company=self.request.user.company).order_by('-created_at')[:10]


class ActivityDashboardView(LoginRequiredMixin, ListView):
    template_name = 'accounts/activity_dashboard.html'
    context_object_name = 'audit_logs'

    def get_queryset(self):
        return AuditLog.objects.filter(company=self.request.user.company).order_by('-timestamp')[:10]


class StatsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/stats_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.user.company

        context.update({
            'company': company,
            'orders_count': Order.objects.filter(company=company).count(),
            'documents_count': Document.objects.filter(company=company).count(),
            'completed_orders': Order.objects.filter(
                company=company,
                status='completed'
            ).count(),
            'new_messages': 3,
        })
        return context


class TeamDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/team_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['company'] = self.request.user.company
        return context

class SupportTicketCreateView(LoginRequiredMixin, CreateView):
    model = SupportTicket
    fields = ['ticket_type', 'message']
    template_name = 'accounts/support_ticket.html'
    success_url = reverse_lazy('accounts:company_dashboard')

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        messages.success(self.request, 'Ваше обращение успешно отправлено!')
        return super().form_valid(form)


def add_delivery_address(request):
    if request.method == 'POST':
        form = DeliveryAddressForm(request.POST)
        if form.is_valid():
            address = form.save(commit=False)
            address.company = request.user.company

            # Если новый адрес помечен как default, снимаем флаг с других адресов
            if address.is_default:
                DeliveryAddress.objects.filter(company=request.user.company).update(is_default=False)

            address.save()

            # Возвращаем JSON ответ
            return JsonResponse({
                'success': True,
                'address_id': address.id,
                'address': address.address,
                'is_default': address.is_default,
                'html': render_to_string('accounts/partials/delivery_address_item.html', {
                    'address': address
                }, request=request)
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            }, status=400)

    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=405)


def delete_delivery_address(request, address_id):
        if request.method == 'DELETE' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                address = DeliveryAddress.objects.get(id=address_id, company=request.user.company)
                address.delete()
                return JsonResponse({'success': True})
            except DeliveryAddress.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Адрес не найден'}, status=404)
        return JsonResponse({'success': False}, status=405)